# coding=utf-8
from __future__ import print_function
import os
import six
from radiomics import setVerbosity
from radiomics.featureextractor import RadiomicsFeatureExtractor
from openpyxl import Workbook
import SimpleITK as sitk
import numpy as np

params = r'E:\Workspace\torch\preprocessing\mine.yaml'
extractor = RadiomicsFeatureExtractor(params)


extractor.addProvenance(provenance_on=False)
extractor.enableAllFeatures()
#extractor.enableFeatureClassByName('shape', False)
# extractor.enableImageTypeByName('lbp', False)
# extractor.enableAllImageTypes()
# extractor.disableAllImageTypes()
# extractor.enableImageTypeByName('LoG', True)

setVerbosity(60)
file = Workbook()
table = file.create_sheet('data')
dataDir = 'I:\case231_cut_xyz'
out_path = 'I:\case231_cut_xyz'
model_name = ['CTA']

label_dir = os.path.join(dataDir,'liver')

row = 1
for index,idx_name in enumerate(sorted(os.listdir(label_dir))):
    print('index: {} index_name: {}'.format(index, idx_name))
    if index >= 0:

        label_case = os.path.join(label_dir, idx_name)

        image_case = label_case.replace('liver','ct').replace('.nii','_0000.nii')

        result = extractor.execute(image_case, label_case)
        column = 1
        for key, val in six.iteritems(result):
            #print('key:{} value{}'.format(key,val))
            assert key is not None
            val = str(val)
            if row == 1:
                table.cell(row=1, column=1, value='case')
                table.cell(row=2, column=1, value=idx_name)
                table.cell(row=row, column=column + 1, value=key)
                table.cell(row=row + 1, column=column + 1, value=val)
            else:
                table.cell(row=row + 1, column=1, value=idx_name)
                table.cell(row=row + 1, column=column + 1, value=val)
            column += 1
    print(row)
    row += 1
    assert len(result) == column - 1

    file.save('{}\\{}.xlsx'.format(out_path, model_name[0]))
